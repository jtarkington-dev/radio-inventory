import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import csv
import tempfile
import os
import subprocess
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime
from database import get_connection

class ReportsWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Radio Reports")
        self.geometry("1000x600")

        self.report_type = tk.StringVar(value="All Radios")
        self.dept_var = tk.StringVar()

        header = tk.Frame(self)
        header.pack(fill=tk.X, padx=10, pady=5)

        left_controls = tk.Frame(header)
        left_controls.pack(side=tk.LEFT, padx=(0, 10))

        right_controls = tk.Frame(header)
        right_controls.pack(side=tk.RIGHT)

        tk.Label(header, text="Report Type:", font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(5, 2))
        self.report_select = ttk.Combobox(header, textvariable=self.report_type, values=[
            "All Radios", "Radios by Department", "Radios in Service", "Disabled Radios", "Missing Radios"
        ], state="readonly", width=25)
        self.report_select.pack(side=tk.LEFT)

        self.dept_combo = ttk.Combobox(header, textvariable=self.dept_var, state="readonly", width=25)
        self.dept_combo.pack(side=tk.LEFT, padx=(10, 0))
        self.dept_combo.pack_forget()

        tk.Button(header, text="Run Report", command=self.run_report).pack(side=tk.LEFT, padx=5)
        tk.Button(header, text="Export Excel", command=self.export_excel).pack(side=tk.LEFT, padx=5)

        self.title_label = tk.Label(self, text="No report selected", font=("Segoe UI", 12, "bold"))
        self.title_label.pack(pady=(5, 0))

        self.tree = ttk.Treeview(self, show="headings")
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.style = ttk.Style()
        self.style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        self.style.configure("Treeview", rowheight=22, font=("Segoe UI", 9))
        self.style.map("Treeview", background=[('selected', '#ececec')])

        self.data_rows = []

    def run_report(self):
        conn = get_connection()
        cursor = conn.cursor()
        report = self.report_type.get()
        self.tree.delete(*self.tree.get_children())

        if report == "All Radios":
            query = '''
                SELECT r.radio_id, r.serial, r.model, d.name, r.assigned_to, r.notes,
                       r.date_received, r.date_issued, r.date_returned,
                       CASE WHEN r.status = 'Active' THEN 'Yes' ELSE 'No' END as in_service
                FROM radios r
                LEFT JOIN departments d ON r.department_id = d.id
            '''
            columns = [
                "Radio ID", "Serial", "Model", "Department", "Assigned To", "Notes",
                "Date Received", "Date Issued", "Date Returned", "In Service"
            ]
            cursor.execute(query)

        elif report == "Radios by Department":
            self.dept_combo.pack(side=tk.LEFT, padx=(10, 0))
            cursor.execute("SELECT name FROM departments")
            dept_names = [row[0] for row in cursor.fetchall()]
            self.dept_combo['values'] = dept_names

            if not self.dept_var.get():
                conn.close()
                return

            cursor.execute('''
                SELECT r.id, r.serial, r.model, r.assigned_to,
                       r.status, r.missing, r.notes
                FROM radios r
                JOIN departments d ON r.department_id = d.id
                WHERE d.name = ?
            ''', (self.dept_var.get(),))
            columns = ["ID", "Serial", "Model", "Assigned", "Status", "Missing", "Notes"]

        elif report == "Radios in Service":
            query = '''
                SELECT r.id, r.serial, r.model, d.name, r.assigned_to
                FROM radios r
                LEFT JOIN departments d ON r.department_id = d.id
                WHERE r.status = 1
            '''
            columns = ["ID", "Serial", "Model", "Department", "Assigned"]
            cursor.execute(query)

        elif report == "Disabled Radios":
            query = '''
                SELECT r.id, r.serial, r.model, d.name, r.assigned_to
                FROM radios r
                LEFT JOIN departments d ON r.department_id = d.id
                WHERE r.status = 0
            '''
            columns = ["ID", "Serial", "Model", "Department", "Assigned"]
            cursor.execute(query)

        elif report == "Missing Radios":
            query = '''
                SELECT r.id, r.serial, r.model, d.name, r.assigned_to
                FROM radios r
                LEFT JOIN departments d ON r.department_id = d.id
                WHERE r.missing = 'Y'
            '''
            columns = ["ID", "Serial", "Model", "Department", "Assigned"]
            cursor.execute(query)

        self.data_rows = cursor.fetchall()
        conn.close()

        self.tree["columns"] = columns
        for col in columns:
            self.tree.heading(col, text=col)
            width = 120
            if col in ("Notes", "Assigned To"):
                width = 180
            self.tree.column(col, anchor="center", width=width)

        for i, row in enumerate(self.data_rows):
            self.tree.insert("", tk.END, values=row, tags=('evenrow' if i % 2 == 0 else 'oddrow'))

        self.tree.tag_configure('evenrow', background='#f8f8f8')
        self.tree.tag_configure('oddrow', background='#e6f2ff')

        self.title_label.config(text=f"{report} - {datetime.datetime.now().strftime('%b %d, %Y %I:%M %p')}")

    def export_excel(self):
        if not self.data_rows:
            messagebox.showwarning("No data", "Run a report first.")
            return

        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Radio Report"

        headers = self.tree["columns"]
        report_title = "GOLDEN NUGGET LAKE CHARLES"
        subtitle = f"REPORT: {self.report_type.get().upper()}   {datetime.datetime.now().strftime('%m/%d/%Y')}"

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = report_title
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Subtitle
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        sub_cell = ws.cell(row=2, column=1)
        sub_cell.value = subtitle
        sub_cell.font = Font(size=12)
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Header
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="305496")
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_index, value=header)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = align
            cell.border = border

        for row_index, row_data in enumerate(self.data_rows, start=5):
            for col_index, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.alignment = align
                cell.border = border

        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 4, 50)

        try:
            wb.save(path)
            messagebox.showinfo("Exported", "Excel file exported successfully.")
        except Exception as e:
            messagebox.showerror("Export Failed", str(e))